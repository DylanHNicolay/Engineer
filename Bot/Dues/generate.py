import discord
from discord import app_commands
from discord.ext import commands
from utils.db import db
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

class GenerateDues(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(name="generate_dues", description="Generate the dues CSV/Excel file.")
    async def generate_dues(self, interaction: discord.Interaction):
        admin_cog = interaction.client.get_cog("Admin")
        if admin_cog is None or not await admin_cog.is_admin(interaction.user):
            await interaction.response.send_message("You do not have permission to use this command.", ephemeral=True)
            return

        await interaction.response.defer(ephemeral=True)

        try:
            # Fetch dues
            dues_record = await db.execute("SELECT * FROM dues LIMIT 1")
            if not dues_record:
                await interaction.followup.send("Dues have not been set. Please use /set_dues_* commands first.")
                return
            dues = dues_record[0]
            starter_dues = dues['starters']
            sub_dues = dues['substitues']
            # non_player_dues = dues['non_player']

            # Fetch active teams
            teams = await db.execute("SELECT * FROM teams WHERE archived = FALSE ORDER BY category_id, team_nick")
            if not teams:
                await interaction.followup.send("No active teams found.")
                return

            # Create Workbook
            wb = openpyxl.Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Styles
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Group teams by category
            teams_by_category = {}
            for team in teams:
                cat_id = team['category_id']
                if cat_id not in teams_by_category:
                    teams_by_category[cat_id] = []
                teams_by_category[cat_id].append(team)

            for cat_id, cat_teams in teams_by_category.items():
                category = interaction.guild.get_channel(cat_id)
                cat_name = category.name if category else f"Category {cat_id}"
                
                # Sanitize sheet name (max 31 chars, no invalid chars)
                sheet_name = "".join(c for c in cat_name if c.isalnum() or c in " -_")[:30]
                if not sheet_name:
                    sheet_name = f"Cat_{cat_id}"
                
                ws = wb.create_sheet(title=sheet_name)
                
                current_row = 1

                for team in cat_teams:
                    team_id = team['team_id']
                    team_nick = team['team_nick']
                    captain_id = team['captain_discord_id']

                    # Fetch members
                    members_records = await db.execute("""
                        SELECT tm.player_discord_id, tm.member_status, p.rcsid 
                        FROM team_members tm
                        LEFT JOIN players p ON tm.player_discord_id = p.player_discord_id
                        WHERE tm.team_id = $1
                    """, team_id)

                    starters_count = sum(1 for m in members_records if m['member_status'] == 'starter')
                    subs_count = sum(1 for m in members_records if m['member_status'] == 'sub')
                    total_players = len(members_records)

                    # ROW 0: Black boxes * 6
                    for col in range(1, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.fill = black_fill
                    current_row += 1

                    # ROW 1
                    # Col A: Team Name:
                    ws.cell(row=current_row, column=1, value="Team Name:")
                    # Col B: [Team Name]
                    ws.cell(row=current_row, column=2, value=team_nick)
                    # Col C: Captain (Yellow)
                    c_cell = ws.cell(row=current_row, column=3, value="Captain")
                    c_cell.fill = yellow_fill
                    # Col D: Red box
                    ws.cell(row=current_row, column=4).fill = red_fill
                    # Col E: Red highlight box
                    ws.cell(row=current_row, column=5).fill = red_fill
                    # Col F: starters = $... (Red highlight box?)
                    f_cell = ws.cell(row=current_row, column=6, value=f"Starters = ${starter_dues}")
                    f_cell.fill = red_fill
                    current_row += 1

                    # ROW 2
                    # Col A: # of players
                    ws.cell(row=current_row, column=1, value="# of players")
                    # Col B: [starters] + [subs]
                    ws.cell(row=current_row, column=2, value=f"{starters_count} + {subs_count}")
                    # Col C: Blank
                    # Col D: Red box
                    ws.cell(row=current_row, column=4).fill = red_fill
                    # Col E: (red box) Subs = $...
                    e_cell = ws.cell(row=current_row, column=5, value=f"Subs = ${sub_dues}")
                    e_cell.fill = red_fill
                    # Col F: Blank?
                    current_row += 1

                    # ROW 3
                    # Col A: League
                    ws.cell(row=current_row, column=1, value="League")
                    # Col E: Disclaimer
                    ws.cell(row=current_row, column=5, value="I understand that I may be charged a $10/25 club fee if I am a member")
                    current_row += 1

                    # ROW 4: Headers (Grey)
                    headers = ["Full Name", "Discord username", "RCSID", "Role", "$", "initials here"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.fill = grey_fill
                        cell.border = thin_border
                    current_row += 1

                    # Player Rows
                    for member_record in members_records:
                        p_id = member_record['player_discord_id']
                        status = member_record['member_status']
                        rcsid = member_record['rcsid'] or ""
                        
                        member = interaction.guild.get_member(p_id)
                        full_name = member.display_name if member else "Unknown"
                        discord_username = member.name if member else "Unknown"
                        
                        # Determine dues for this player
                        player_due = starter_dues if status == 'starter' else sub_dues

                        row_values = [full_name, discord_username, rcsid, status, f"${player_due}", ""]
                        
                        for col, val in enumerate(row_values, 1):
                            cell = ws.cell(row=current_row, column=col, value=val)
                            cell.border = thin_border
                            
                            # Highlight captain
                            if p_id == captain_id:
                                cell.fill = yellow_fill
                        
                        current_row += 1
                    
                    # Add some spacing between teams
                    current_row += 2

                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            await interaction.followup.send(file=discord.File(buffer, filename="dues.xlsx"))

        except Exception as e:
            await interaction.followup.send(f"An error occurred: {e}")
            import traceback
            traceback.print_exc()

async def setup(bot: commands.Bot):
    await bot.add_cog(GenerateDues(bot))
